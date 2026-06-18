#!/usr/bin/env python3
"""Demo: parse catalog site → Excel/CSV. Portfolio piece for Kwork."""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


BASE_URL = "http://books.toscrape.com/"
USER_AGENT = "Mozilla/5.0 (portfolio-demo; +python-requests)"


@dataclass
class Product:
    title: str
    price_gbp: float
    availability: str
    rating: str
    category: str
    url: str


def fetch_html(url: str) -> str:
    resp = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=20)
    resp.raise_for_status()
    return resp.text


def parse_price(text: str) -> float:
    match = re.search(r"[\d.]+", text.replace(",", ""))
    return float(match.group()) if match else 0.0


def parse_page(html: str, category: str, page_url: str) -> list[Product]:
    soup = BeautifulSoup(html, "html.parser")
    items: list[Product] = []
    for article in soup.select("article.product_pod"):
        title = article.select_one("h3 a")["title"].strip()
        link = urljoin(page_url, article.select_one("h3 a")["href"])
        price = parse_price(article.select_one("p.price_color").get_text(strip=True))
        availability = article.select_one("p.instock.availability").get_text(" ", strip=True)
        rating = next(
            (cls for cls in article.select_one("p.star-rating")["class"] if cls != "star-rating"),
            "?",
        )
        items.append(
            Product(
                title=title,
                price_gbp=price,
                availability=availability,
                rating=rating,
                category=category,
                url=link,
            )
        )
    return items


def collect_products(max_pages: int = 3) -> list[Product]:
    products: list[Product] = []
    index_html = fetch_html(BASE_URL)
    index_soup = BeautifulSoup(index_html, "html.parser")
    categories = [
        ("All", BASE_URL),
        *[
            (a.get_text(strip=True), urljoin(BASE_URL, a["href"]))
            for a in index_soup.select("div.side_categories ul li ul li a")
        ][:4],
    ]
    for category, cat_url in categories:
        page_url = cat_url
        for _ in range(max_pages):
            html = fetch_html(page_url)
            products.extend(parse_page(html, category, page_url))
            soup = BeautifulSoup(html, "html.parser")
            next_link = soup.select_one("li.next a")
            if not next_link:
                break
            page_url = urljoin(page_url, next_link["href"])
    return products


def save_csv(products: list[Product], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = ["title,price_gbp,availability,rating,category,url"]
    for p in products:
        row = [
            p.title.replace('"', '""'),
            f"{p.price_gbp:.2f}",
            p.availability.replace('"', '""'),
            p.rating,
            p.category.replace('"', '""'),
            p.url,
        ]
        lines.append(",".join(f'"{v}"' if i in (0, 2, 4) else v for i, v in enumerate(row)))
    path.write_text("\n".join(lines), encoding="utf-8")


def save_xlsx(products: list[Product], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    headers = ["Название", "Цена GBP", "Наличие", "Рейтинг", "Категория", "Ссылка"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for col, cell in enumerate(ws[1], start=1):
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for p in products:
        ws.append([p.title, p.price_gbp, p.availability, p.rating, p.category, p.url])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        row[5].hyperlink = row[5].value
        row[5].style = "Hyperlink"
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 48
    meta = wb.create_sheet("Meta")
    meta.append(["Сгенерировано", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")])
    meta.append(["Источник", BASE_URL])
    meta.append(["Записей", len(products)])
    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Catalog parser demo → Excel/CSV")
    parser.add_argument("--pages", type=int, default=2, help="Pages per category")
    parser.add_argument(
        "--out-dir",
        type=Path,
        default=Path(__file__).resolve().parents[2] / "samples",
    )
    args = parser.parse_args()
    products = collect_products(max_pages=args.pages)
    if not products:
        print("No products parsed", file=sys.stderr)
        return 1
    out_dir = args.out_dir
    csv_path = out_dir / "catalog_products.csv"
    xlsx_path = out_dir / "catalog_products.xlsx"
    save_csv(products, csv_path)
    save_xlsx(products, xlsx_path)
    print(f"OK: {len(products)} products")
    print(f"CSV:  {csv_path}")
    print(f"XLSX: {xlsx_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
